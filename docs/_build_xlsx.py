"""Build a brand-styled experiment matrix scorecard xlsx for an autoresearch session.

Reads:
  ~/.gstack/projects/<slug>/autoresearch/state.json

Writes (default):
  results/<date>_<scope>/scorecard_<date>.xlsx

Generates up to 5 sheets (gracefully skipped when state data is absent):
  Sheet 1 — Axis Matrix:       hyperparameter scorecard, params × target
  Sheet 2 — Per-task headline: best config + key metrics per task/target
  Sheet 3 — HPO detail:        hyperparameter sweep results table
  Sheet 4 — Future directions: deferred / next-iteration items
  Sheet 5 — Legend:            color & status reference

Brand palette from skills/_shared/branding.py:
  header bg:  INK   #141414, white text, bold
  sub-header: dark  #222222, light-grey text
  winner:     TURQUOISE #40E0D0, ink text, bold (no glyph)
  section:    TURQUOISE fill on section-header rows
  warning:    AMBER #F0C840, white text
  deferred:   light grey #E8E8E8, dark text
  body:       PAPER #FAFAFA fill, ink text
  tab color:  TURQUOISE on every sheet

Run:
    python skills/autoresearch/templates/_build_xlsx.py \\
        --date 2026-04-30 --scope fw-arch-sweep

    # explicit output path:
    python skills/autoresearch/templates/_build_xlsx.py \\
        --date 2026-04-30 --scope fw-arch-sweep \\
        --output /path/to/scorecard.xlsx
"""
from __future__ import annotations

import argparse
import json
import os
import re
import sys
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import (
        Alignment,
        Border,
        Font,
        GradientFill,
        PatternFill,
        Side,
    )
    from openpyxl.utils import get_column_letter
except ImportError:
    raise SystemExit(
        "openpyxl not installed. Run: pip install openpyxl"
    )

# ---------------------------------------------------------------------------
# Import shared styling primitives from skills/_shared/branding_xlsx.py.
#
# Resolution order:
#   1. $SUPERSTACK_HOME/skills/_shared        (explicit override)
#   2. ../../_shared (relative to this file)  (works in-repo at templates/)
#   3. ~/arcadia/superstack/skills/_shared    (default install location)
#   4. ~/.claude/skills/_shared               (local-skills install location)
#
# We need the absolute-path variants because this template is dropped into
# projects under <project>/docs/_build_xlsx.py at init time, where the
# parents[2] heuristic resolves to <project>/ and breaks the import.
# ---------------------------------------------------------------------------

_HERE = Path(__file__).resolve()
_SHARED_CANDIDATES = [
    Path(os.environ.get("SUPERSTACK_HOME", "")) / "skills" / "_shared"
        if os.environ.get("SUPERSTACK_HOME") else None,
    _HERE.parents[2] / "_shared",
    Path.home() / "arcadia" / "superstack" / "skills" / "_shared",
    Path.home() / ".claude" / "skills" / "_shared",
]
_SHARED = next((p for p in _SHARED_CANDIDATES if p and p.is_dir()), None)
if _SHARED is None:
    raise SystemExit(
        "Could not locate skills/_shared/. Set $SUPERSTACK_HOME or install superstack."
    )
sys.path.insert(0, str(_SHARED))

try:
    from branding_xlsx import (  # type: ignore
        TURQUOISE as _TURQUOISE,
        DEEPPINK  as _DEEPPINK,
        AMBER     as _AMBER,
        BLUEVIOLET as _BLUEVIOLET,
        INK       as _INK,
        PAPER     as _PAPER,
        WHITE     as _WHITE,
        MUTED     as _MUTED,
        DEFERRED_GREY as _DEFERRED_GREY,
        DARK_SUBHEADER as _DARK_SUBHEADER,
        LIGHT_GREY_TXT as _LIGHT_GREY_TXT,
        RAN_LIGHT as _RAN_LIGHT,
        _fill,
        _font,
        _align,
        _write_header_row,
        _write_body_row,
        _write_winner_row,
        _write_deferred_row,
        _write_warning_row,
        _set_tab_color,
        _set_col_widths,
        _freeze_header as _freeze,
        add_glyph_to_cell as _add_glyph,
    )
except ImportError:
    # Fallback: inline definitions so autoresearch keeps working even if
    # the shared module is unavailable (e.g. isolated testing environments).
    from branding import (  # type: ignore
        TURQUOISE, DEEPPINK, AMBER, BLUEVIOLET, INK, PAPER, WHITE, MUTED
    )
    _TURQUOISE  = TURQUOISE.lstrip("#")
    _DEEPPINK   = DEEPPINK.lstrip("#")
    _AMBER      = AMBER.lstrip("#")
    _BLUEVIOLET = BLUEVIOLET.lstrip("#")
    _INK        = INK.lstrip("#")
    _PAPER      = PAPER.lstrip("#")
    _WHITE      = WHITE.lstrip("#")
    _MUTED      = MUTED.lstrip("#")
    _DEFERRED_GREY  = "E8E8E8"
    _DARK_SUBHEADER = "222222"
    _LIGHT_GREY_TXT = "CCCCCC"
    _RAN_LIGHT      = "CDEFEB"
    _SOLID = "solid"

    def _fill(hex_color):
        return PatternFill(fill_type=_SOLID, fgColor=hex_color)

    def _font(hex_color=_INK, *, bold=False, italic=False, size=10, name="Geist"):
        return Font(color=hex_color, bold=bold, italic=italic, size=size, name=name)

    def _align(horizontal="left", vertical="center", wrap=False):
        return Alignment(horizontal=horizontal, vertical=vertical, wrap_text=wrap)

    def _write_header_row(ws, row_num, values, *, col_offset=1):
        for i, val in enumerate(values):
            c = ws.cell(row=row_num, column=col_offset + i, value=val)
            c.fill = _fill(_INK)
            c.font = _font(_WHITE, bold=True)
            c.alignment = _align()

    def _write_body_row(ws, row_num, values, *, col_offset=1, alt=False):
        for i, val in enumerate(values):
            c = ws.cell(row=row_num, column=col_offset + i, value=val)
            c.fill = _fill(_PAPER)
            c.font = _font(_INK)
            c.alignment = _align(wrap=True)

    def _write_winner_row(ws, row_num, values, *, col_offset=1):
        for i, val in enumerate(values):
            c = ws.cell(row=row_num, column=col_offset + i, value=val)
            c.fill = _fill(_TURQUOISE)
            c.font = _font(_INK, bold=True)
            c.alignment = _align()

    def _write_deferred_row(ws, row_num, values, *, col_offset=1):
        for i, val in enumerate(values):
            c = ws.cell(row=row_num, column=col_offset + i, value=val)
            c.fill = _fill(_DEFERRED_GREY)
            c.font = _font(_INK)
            c.alignment = _align(wrap=True)

    def _write_warning_row(ws, row_num, values, *, col_offset=1):
        for i, val in enumerate(values):
            c = ws.cell(row=row_num, column=col_offset + i, value=val)
            c.fill = _fill(_AMBER)
            c.font = _font(_INK, bold=True)
            c.alignment = _align(wrap=True)

    def _set_tab_color(ws, color_hex=_TURQUOISE):
        ws.sheet_properties.tabColor = color_hex

    def _set_col_widths(ws, widths):
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def _freeze(ws, cell="A2"):
        ws.freeze_panes = cell


# ---------------------------------------------------------------------------
# Scorecard-specific helpers not in the shared module
# ---------------------------------------------------------------------------

def _write_subheader_row(ws, row_num: int, value: str, n_cols: int) -> None:
    """Dark #222 sub-header row spanning n_cols, light-grey text."""
    c = ws.cell(row=row_num, column=1, value=value)
    c.fill = _fill(_DARK_SUBHEADER)
    c.font = _font(_LIGHT_GREY_TXT, size=9)
    c.alignment = _align()
    for col in range(2, n_cols + 1):
        ws.cell(row=row_num, column=col).fill = _fill(_DARK_SUBHEADER)


def _write_section_row(ws, row_num: int, value: str, n_cols: int) -> None:
    """Turquoise section header row (like 'Backbone', 'Inputs')."""
    c = ws.cell(row=row_num, column=1, value=value)
    c.fill = _fill(_TURQUOISE)
    c.font = _font(_INK, bold=True)
    c.alignment = _align()
    for col in range(2, n_cols + 1):
        ws.cell(row=row_num, column=col).fill = _fill(_TURQUOISE)


def _write_metric_cell(ws, row_num: int, col_num: int, value: str) -> None:
    """Blueviolet metric value cell."""
    c = ws.cell(row=row_num, column=col_num, value=value)
    c.fill = _fill(_PAPER)
    c.font = _font(_BLUEVIOLET, bold=True, name="Geist Mono")
    c.alignment = _align(horizontal="center")


# ---------------------------------------------------------------------------
# State reading
# ---------------------------------------------------------------------------

def _read_state(scope: str) -> dict:
    """Locate state.json for the current project + verify scope matches.

    autoresearch keys state.json by *project slug* (derived from cwd), not by
    *scope slug*. Search order:
      1. ~/.gstack/projects/<cwd-basename>/autoresearch/state.json
      2. Fallback: ~/.gstack/projects/<scope>/autoresearch/state.json
         (covers the legacy case where someone seeded state by scope slug).

    Once located, verify state["scope_slug"] == scope to fail loud on mismatch.
    """
    gstack_home = Path(os.environ.get("GSTACK_HOME", Path.home() / ".gstack"))
    cwd_slug = Path.cwd().resolve().name
    candidates = [
        gstack_home / "projects" / cwd_slug / "autoresearch" / "state.json",
        gstack_home / "projects" / scope    / "autoresearch" / "state.json",
    ]
    state_path = next((p for p in candidates if p.exists()), None)
    if state_path is None:
        raise SystemExit(
            f"No state.json found. Looked in:\n  "
            + "\n  ".join(str(p) for p in candidates)
        )
    with state_path.open() as fh:
        state = json.load(fh)
    state_scope = state.get("scope_slug", "")
    if state_scope and state_scope != scope:
        print(
            f"warning: --scope='{scope}' but state.json has scope_slug='{state_scope}'. "
            f"Using state.json from {state_path}.",
            file=sys.stderr,
        )
    return state


def _humanize(slug: str) -> str:
    if not slug:
        return ""
    slug = re.sub(r"^iter-\d+_", "", slug)
    parts = re.split(r"[-_]+", slug)
    return " ".join(p[:1].upper() + p[1:] if p else "" for p in parts).strip()


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def _build_per_task(wb: Workbook, state: dict) -> None:
    """Sheet 2 — Per-task headline: best config + metrics per task/target."""
    results_history: list[dict] = state.get("results_history", [])
    target_metric: str = state.get("target_metric", "") or ""

    tasks: dict[str, list[dict]] = {}
    for r in results_history:
        task = r.get("axes", {}).get("task") or r.get("task") or ""
        if not task:
            continue
        tasks.setdefault(task, []).append(r)

    if not tasks:
        return

    ws = wb.create_sheet("Per-task headline")
    _set_tab_color(ws)

    scope_slug = state.get("scope_slug", "")
    sub = state.get("scope", "") or _humanize(scope_slug)

    c = ws.cell(row=1, column=1, value=f"Per-task headline — {_humanize(scope_slug)}")
    c.fill = _fill(_INK)
    c.font = _font(_WHITE, bold=True, size=12)
    for col in range(2, 8):
        ws.cell(row=1, column=col).fill = _fill(_INK)

    _write_subheader_row(ws, 2, sub, 7)

    hdrs = ["Task", "Best config", target_metric or "Metric", "Status", "Iterations", "Best axes", "Notes"]
    for i, h in enumerate(hdrs):
        c = ws.cell(row=3, column=i + 1, value=h)
        c.fill = _fill(_TURQUOISE)
        c.font = _font(_INK, bold=True)
        c.alignment = _align()

    row_num = 4
    for task, task_results in tasks.items():
        completed = [r for r in task_results if r.get("status") == "complete"]
        if not completed:
            completed = task_results
        best = max(completed, key=lambda r: r.get("metric_value") or 0, default=None)
        if not best:
            continue

        mv = best.get("metric_value", "—")
        axes_str = ", ".join(f"{k}={v}" for k, v in (best.get("axes") or {}).items())
        candidate_id = best.get("id", "")
        fix_attempts = best.get("fix_attempts", 0)
        notes = best.get("notes", "") or ""

        row_vals = [
            _humanize(str(task)),
            _humanize(candidate_id),
            str(mv) if mv is not None else "—",
            best.get("status", "—"),
            str(len(task_results)),
            axes_str,
            notes,
        ]
        _write_body_row(ws, row_num, row_vals)
        _write_metric_cell(ws, row_num, 3, str(mv) if mv is not None else "—")
        row_num += 1

    _set_col_widths(ws, [18, 22, 12, 12, 10, 30, 30])
    _freeze(ws)


def _build_hpo_detail(wb: Workbook, state: dict) -> None:
    """Sheet 3 — HPO detail: hyperparameter sweep results.

    Expects state.json to carry a 'hpo_runs' key:
      [{"id": "...", "params": {...}, "metrics": {...}}, ...]
    Gracefully skipped if absent or empty.
    """
    hpo_runs: list[dict] = state.get("hpo_runs") or []
    if not hpo_runs:
        return

    ws = wb.create_sheet("HPO detail")
    _set_tab_color(ws)

    scope_slug = state.get("scope_slug", "")
    target_metric: str = state.get("target_metric", "") or ""

    c = ws.cell(row=1, column=1, value=f"HPO detail — {_humanize(scope_slug)}")
    c.fill = _fill(_INK)
    c.font = _font(_WHITE, bold=True, size=12)

    param_keys: list[str] = []
    metric_keys: list[str] = []
    seen_p: set[str] = set()
    seen_m: set[str] = set()
    for run in hpo_runs:
        for k in (run.get("params") or {}).keys():
            if k not in seen_p:
                param_keys.append(k)
                seen_p.add(k)
        for k in (run.get("metrics") or {}).keys():
            if k not in seen_m:
                metric_keys.append(k)
                seen_m.add(k)

    all_cols = ["Run"] + param_keys + metric_keys
    n_cols = len(all_cols)

    for col in range(2, n_cols + 1):
        ws.cell(row=1, column=col).fill = _fill(_INK)

    _write_subheader_row(ws, 2, state.get("scope", "") or "", n_cols)

    for i, h in enumerate(all_cols):
        c = ws.cell(row=3, column=i + 1, value=h)
        c.fill = _fill(_TURQUOISE)
        c.font = _font(_INK, bold=True)
        c.alignment = _align(horizontal="center")

    best_metric_vals: dict[str, float] = {}
    for mk in metric_keys:
        vals = [run.get("metrics", {}).get(mk) for run in hpo_runs
                if run.get("metrics", {}).get(mk) is not None]
        numeric = [v for v in vals if isinstance(v, (int, float))]
        if numeric:
            best_metric_vals[mk] = max(numeric)

    row_num = 4
    for run in hpo_runs:
        run_id = run.get("id", f"run-{row_num - 3}")
        params = run.get("params") or {}
        metrics = run.get("metrics") or {}

        row_vals = [run_id] + [str(params.get(k, "—")) for k in param_keys]
        c_start = len(row_vals) + 1

        _write_body_row(ws, row_num, row_vals)

        for mi, mk in enumerate(metric_keys):
            mv = metrics.get(mk)
            col_num = c_start + mi
            if mv is None:
                ws.cell(row=row_num, column=col_num, value="—")
            else:
                is_best = (isinstance(mv, (int, float))
                           and best_metric_vals.get(mk) == mv)
                c = ws.cell(row=row_num, column=col_num, value=mv)
                if is_best:
                    c.fill = _fill(_TURQUOISE)
                    c.font = _font(_INK, bold=True)
                    c.value = f"{mv} ★"
                else:
                    c.fill = _fill(_PAPER)
                    c.font = _font(_BLUEVIOLET, bold=False, name="Geist Mono")
                c.alignment = _align(horizontal="center")

        row_num += 1

    param_widths = [14] + [14] * len(param_keys) + [14] * len(metric_keys)
    _set_col_widths(ws, param_widths)
    _freeze(ws)


def _build_future_directions(wb: Workbook, state: dict) -> None:
    """Sheet 4 — Future directions: deferred / next-iteration items.

    Expects state.json to carry a 'future_directions' key:
      [{"item": "...", "effort": "S/M/L", "expected_lift": "...",
        "notes": "...", "priority": "high/normal", "deferred": true/false}, ...]
    Or falls back to candidate_queue items with status=pending that look like
    deferred work.

    Gracefully skipped if no data.
    """
    future: list[dict] = state.get("future_directions") or []

    if not future:
        for c in state.get("candidate_queue") or []:
            if c.get("status") == "pending":
                axes_str = ", ".join(
                    f"{k}={v}" for k, v in (c.get("axes") or {}).items()
                )
                future.append({
                    "item": axes_str or c.get("id", ""),
                    "effort": "?",
                    "expected_lift": "",
                    "notes": "",
                    "deferred": True,
                })

    if not future:
        return

    ws = wb.create_sheet("Future directions")
    _set_tab_color(ws)

    scope_slug = state.get("scope_slug", "")

    c = ws.cell(row=1, column=1, value=f"Future directions — {_humanize(scope_slug)}")
    c.fill = _fill(_INK)
    c.font = _font(_WHITE, bold=True, size=12)
    for col in range(2, 6):
        ws.cell(row=1, column=col).fill = _fill(_INK)

    _write_subheader_row(ws, 2, state.get("scope", "") or "", 5)

    _write_header_row(ws, 3, ["#", "Experiment", "Effort", "Expected lift / value", "Notes / unlock"])

    row_num = 4
    for i, item in enumerate(future, start=1):
        label = item.get("item", "")
        effort = item.get("effort", "?")
        lift = item.get("expected_lift", "")
        notes = item.get("notes", "")
        deferred = item.get("deferred", True)
        priority = item.get("priority", "normal")

        row_vals = [str(i), label, effort, lift, notes]

        if priority == "high":
            _write_warning_row(ws, row_num, row_vals)
        elif deferred:
            _write_deferred_row(ws, row_num, row_vals)
        else:
            _write_body_row(ws, row_num, row_vals)

        row_num += 1

    _set_col_widths(ws, [4, 40, 8, 30, 30])
    _freeze(ws)


def _build_legend(wb: Workbook) -> None:
    """Sheet 5 — Legend: color and status glyph reference."""
    ws = wb.create_sheet("Legend")
    _set_tab_color(ws)

    c = ws.cell(row=1, column=1, value="Legend — color & status reference")
    c.fill = _fill(_INK)
    c.font = _font(_WHITE, bold=True, size=12)
    for col in range(2, 5):
        ws.cell(row=1, column=col).fill = _fill(_INK)

    ws.cell(row=2, column=1).value = ""

    row = 3
    c = ws.cell(row=row, column=1, value="Status codes")
    c.fill = _fill(_DEEPPINK)
    c.font = _font(_WHITE, bold=True)
    for col in range(2, 5):
        ws.cell(row=row, column=col).fill = _fill(_DEEPPINK)

    row += 1
    _write_header_row(ws, row, ["", "Status", "Meaning", "Sample fill"])

    statuses = [
        ("WON",      "Ran this iteration and was the winner for its experiment",     "★ WON",       _TURQUOISE),
        ("RAN",      "Ran this iteration; informative but not the top choice",       "✓ RAN",       _RAN_LIGHT),
        ("LOST",     "Ran and was ACTIVELY HARMFUL — avoid in future",               "✗ LOST",      _DEEPPINK),
        ("NEW",      "Added mid-session; added on top of the original plan",         "◆ NEW",       _AMBER),
        ("DEF",      "Matrix item was never attempted (deferred to next sprint)",    "▶ DEFERRED",  _DEFERRED_GREY),
        ("HALTED",   "Session halted due to infra failure",                          "◼ HALTED",    _MUTED),
    ]

    for stat, meaning, sample, bg_hex in statuses:
        row += 1
        ws.cell(row=row, column=1).fill = _fill(_PAPER)
        c = ws.cell(row=row, column=2, value=stat)
        c.fill = _fill(_PAPER)
        c.font = _font(_INK)
        c = ws.cell(row=row, column=3, value=meaning)
        c.fill = _fill(_PAPER)
        c.font = _font(_INK)
        c.alignment = _align(wrap=True)
        c = ws.cell(row=row, column=4, value=sample)
        c.fill = _fill(bg_hex)
        c.font = _font(_INK, bold=(bg_hex in (_TURQUOISE, _AMBER, _DEEPPINK)))
        c.alignment = _align(horizontal="center")

    row += 2

    c = ws.cell(row=row, column=1, value="Color palette — priority order")
    c.fill = _fill(_DEEPPINK)
    c.font = _font(_WHITE, bold=True)
    for col in range(2, 5):
        ws.cell(row=row, column=col).fill = _fill(_DEEPPINK)

    row += 1
    palette = [
        ("#1", "Turquoise",  "headline metric · winner callouts · section headers",    _TURQUOISE),
        ("#2", "Deeppink",   "secondary accents · comparison points · LOST status",    _DEEPPINK),
        ("#3", "Amber",      "tertiary — NEW status · warning / high-priority rows",   _AMBER),
        ("#4", "Blueviolet", "metric values · quaternary callouts",                    _BLUEVIOLET),
        ("#5", "Paper",      "standard body cell fill",                                _PAPER),
        ("#6", "Ink",        "header / title backgrounds",                             _INK),
        ("#7", "Deferred",   "deferred / pending items",                               _DEFERRED_GREY),
    ]
    for rank, name, desc, bg_hex in palette:
        c = ws.cell(row=row, column=1, value=rank)
        c.fill = _fill(_PAPER)
        c.font = _font(_MUTED)
        c = ws.cell(row=row, column=2, value=name)
        c.fill = _fill(_PAPER)
        c.font = _font(_INK)
        c = ws.cell(row=row, column=3, value=desc)
        c.fill = _fill(_PAPER)
        c.font = _font(_INK)
        c.alignment = _align(wrap=True)
        c = ws.cell(row=row, column=4, value=f"#{bg_hex}")
        c.fill = _fill(bg_hex)
        c.font = _font(_WHITE if bg_hex in (_INK, _BLUEVIOLET, _DEEPPINK) else _INK)
        c.alignment = _align(horizontal="center")
        row += 1

    row += 1
    c = ws.cell(row=row, column=1,
                value="Rebuild: python skills/autoresearch/templates/_build_xlsx.py --scope <slug>")
    c.fill = _fill(_INK)
    c.font = _font(_MUTED, size=9)
    for col in range(2, 5):
        ws.cell(row=row, column=col).fill = _fill(_INK)

    _set_col_widths(ws, [6, 14, 48, 20])
    _freeze(ws, "A2")


def _build_axis_matrix(wb: Workbook, state: dict, date_str: str) -> None:
    """Hyperparameter scorecard (CurieDx style).

    Renders one Axis Matrix sheet per declared target axis. ``state.target_axis``
    may be:

    - a **string** (e.g. ``"pathogen"``) — single sheet
    - a **list of strings** (e.g. ``["pathogen", "tissue"]``) — one sheet per
      target axis, each rendering that axis as columns and treating every
      remaining axis (including the OTHER target axes) as row sections
    - ``null`` or absent — fallback to common names (``target``, ``gene``,
      ``pathogen``, ``disease``, ``label``, ``task``); single sheet using the
      first match. If none match, renders one sheet with a single
      ``Best metric`` column.

    Each per-target sheet computes per-(parameter, target_value) best metric
    by aggregating across all OTHER axes (so each sheet is a marginal view).

    See :func:`_build_one_axis_matrix` for the per-sheet layout details.
    """
    axes: dict = state.get("axes", {})
    if not axes:
        return

    declared_target = state.get("target_axis")
    fallback_target_keys = (
        "target", "gene", "pathogen", "disease", "label", "task",
    )

    # Resolve to a list of target axis keys.
    target_keys: list = []
    if isinstance(declared_target, list):
        target_keys = [t for t in declared_target if t in axes]
    elif isinstance(declared_target, str) and declared_target in axes:
        target_keys = [declared_target]
    else:
        fallback = next(
            (k for k in fallback_target_keys if k in axes), None
        )
        if fallback is not None:
            target_keys = [fallback]

    if not target_keys:
        # No target axis at all: render a single sheet with a "Best metric" col.
        _build_one_axis_matrix(
            wb, state, date_str,
            target_key=None,
            sheet_name="Axis Matrix",
        )
        return

    multi_sheet = len(target_keys) > 1
    for tk in target_keys:
        sheet_name = f"Axis Matrix — {tk}" if multi_sheet else "Axis Matrix"
        _build_one_axis_matrix(
            wb, state, date_str,
            target_key=tk,
            sheet_name=sheet_name,
        )


def _build_one_axis_matrix(
    wb: Workbook,
    state: dict,
    date_str: str,
    *,
    target_key: str | None,
    sheet_name: str,
) -> None:
    """Render a single Axis Matrix sheet.

    Layout (multi-target-value mode, e.g. target_key="pathogen")::

        | Parameter        | flu     | covid | strep | RSV   |
        | ─ Method ─       | (turquoise stripe across all cols)|
        | xgb              |  0.849  | 0.74  | 0.796 | 0.717 |  ← per-target-value
        | tabpfn           |  0.81   | 0.808 | 0.71  | 0.66  |    winners highlighted
        | ─ Feature Mode ─ | (deeppink stripe)                 |    in that col
        | mi_top20         |  0.849  | 0.78  | 0.796 | 0.717 |
        | ...

    Section-header rows cycle through the brand palette
    (turquoise → deeppink → amber → blueviolet). Body cells use no
    fill so default Excel gridlines remain visible. Per-axis winners
    are derived from ``results_history`` (max metric per axis-value
    per target value, complete runs only) — when multiple target axes
    exist, this aggregates across the others (a marginal view).

    Hyperparameter axis order = dict insertion order in ``state.axes``,
    which by convention reflects "tested first → tested last."
    """
    axes: dict = state.get("axes", {})
    if not axes:
        return
    results_history: list[dict] = state.get("results_history", [])
    scope_slug = state.get("scope_slug", "")

    target_values: list = []
    if target_key is not None and target_key in axes:
        tv = axes[target_key]
        target_values = list(tv) if isinstance(tv, list) else [tv]

    # Hyperparameter axes (everything except THIS target axis).
    # Note: other target axes (if any) are still rendered as row sections —
    # this makes the per-target-axis sheets a marginal view across the others.
    axis_items = [
        (name, vals if isinstance(vals, list) else [vals])
        for name, vals in axes.items()
        if name != target_key
    ]
    if not axis_items or all(len(v) == 0 for _, v in axis_items):
        return

    # Column structure: A=Parameter, then one per target value (or single metric col).
    multi_target = len(target_values) > 1
    metric_cols = target_values if multi_target else [None]
    n_cols = 1 + len(metric_cols)

    ws = wb.create_sheet(sheet_name)
    _set_tab_color(ws)

    # Title bar
    title = f"{_humanize(scope_slug)} — {sheet_name} · {date_str}"
    c = ws.cell(row=1, column=1, value=title)
    c.fill = _fill(_INK)
    c.font = _font(_WHITE, bold=True, size=12)
    c.alignment = _align()
    for col in range(2, n_cols + 1):
        ws.cell(row=1, column=col).fill = _fill(_INK)

    sub = state.get("scope", "") or _humanize(scope_slug)
    _write_subheader_row(ws, 2, sub, n_cols)

    # Header row
    target_metric = state.get("target_metric") or "Best metric"
    if multi_target:
        headers = ["Parameter"] + [str(g) for g in target_values]
    else:
        headers = ["Parameter", _humanize(target_metric)]
    _write_header_row(ws, 4, headers)

    # Cycle of accent colors for axis sections
    accent_cycle = [_TURQUOISE, _DEEPPINK, _AMBER, _BLUEVIOLET]
    accent_white_text = {_DEEPPINK, _BLUEVIOLET}  # contrast on dark fills

    def _best_metric(ax_name, ax_value, target_value):
        """Best metric_value across complete runs matching the filters."""
        matches = []
        for r in results_history:
            r_axes = r.get("axes", {}) or {}
            if r.get("status") != "complete":
                continue
            if r_axes.get(ax_name) != ax_value:
                continue
            if target_value is not None and r_axes.get(target_key) != target_value:
                continue
            mv = r.get("metric_value")
            if mv is None:
                continue
            matches.append(mv)
        return max(matches) if matches else None

    row = 5
    for ax_idx, (ax_name, vals) in enumerate(axis_items):
        accent = accent_cycle[ax_idx % len(accent_cycle)]
        text_on_accent = _WHITE if accent in accent_white_text else _INK

        # Section-header row spanning all columns (◆ bullet prefix).
        c = ws.cell(row=row, column=1, value=f"◆  {_humanize(ax_name)}")
        c.fill = _fill(accent)
        c.font = _font(text_on_accent, bold=True, size=11)
        c.alignment = _align()
        for col in range(2, n_cols + 1):
            ws.cell(row=row, column=col).fill = _fill(accent)
        row += 1

        # Pre-compute the best metric per (target, ax_value) so we can find
        # the winning ax_value per target column for this section.
        section_metrics: dict = {}  # (target_value, ax_value) -> mv
        for v in vals:
            for target_value in metric_cols:
                section_metrics[(target_value, v)] = _best_metric(
                    ax_name, v, target_value
                )

        # Per-target-column winning ax_value (the one with the max metric).
        winning_value_per_col: dict = {}  # target_value -> ax_value
        for target_value in metric_cols:
            best_v = None
            best_mv = None
            for v in vals:
                mv = section_metrics.get((target_value, v))
                if mv is None:
                    continue
                if best_mv is None or mv > best_mv:
                    best_mv = mv
                    best_v = v
            if best_v is not None:
                winning_value_per_col[target_value] = best_v

        # Body rows
        for v in vals:
            # Param cell — no fill so gridlines show through.
            pcell = ws.cell(row=row, column=1, value=str(v))
            pcell.font = _font(_INK)
            pcell.alignment = _align()

            for col_idx, target_value in enumerate(metric_cols, start=2):
                mv = section_metrics.get((target_value, v))
                is_winner = (
                    mv is not None
                    and winning_value_per_col.get(target_value) == v
                )
                if mv is None:
                    metric_str = "▶ DEFERRED"
                elif is_winner:
                    metric_str = f"{mv} ★"  # solid unicode star (text glyph)
                else:
                    metric_str = f"{mv}"
                mcell = ws.cell(row=row, column=col_idx, value=metric_str)
                mcell.alignment = _align(horizontal="center")
                if mv is None:
                    mcell.fill = _fill(_DEFERRED_GREY)
                    mcell.font = _font(_MUTED, italic=True, size=9)
                elif is_winner:
                    mcell.fill = _fill(accent)
                    mcell.font = _font(text_on_accent, bold=True,
                                       name="Geist Mono")
                else:
                    mcell.font = _font(_BLUEVIOLET, name="Geist Mono")
            row += 1

    # Column widths: param col wider, metric cols narrower in multi-target mode.
    if multi_target:
        col_widths = [28] + [16] * len(target_values)
    else:
        col_widths = [32, 18]
    _set_col_widths(ws, col_widths)
    _freeze(ws, "B5")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--date",
        default=datetime.now().strftime("%Y-%m-%d"),
        help="Session date (YYYY-MM-DD). Default: today.",
    )
    parser.add_argument("--scope", required=True, help="Session scope slug.")
    parser.add_argument(
        "--output",
        default=None,
        help=(
            "Output path for the xlsx. "
            "Default: results/<date>_<scope>/scorecard_<date>.xlsx"
        ),
    )
    args = parser.parse_args()

    state = _read_state(args.scope)

    path_date = args.date.replace("-", "")
    if args.output:
        out = Path(args.output)
    else:
        out_dir = Path("results") / f"{path_date}_{args.scope}"
        out_dir.mkdir(parents=True, exist_ok=True)
        out = out_dir / f"scorecard_{path_date}.xlsx"

    out.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    _build_axis_matrix(wb, state, args.date)
    _build_per_task(wb, state)
    _build_hpo_detail(wb, state)
    _build_future_directions(wb, state)
    _build_legend(wb)

    wb.save(out)
    print(f"wrote {out}")


if __name__ == "__main__":
    main()
